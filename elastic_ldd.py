import os
import csv
import sys
import argparse
import json
import time
from copy import deepcopy
from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl import Workbook
# from openpyxl.compat import range
from elasticsearch import Elasticsearch

# Constants
HEADERS = []
ATTRIBUTES = []
DATA = []
# Ignore in indexing
AFIELDS = ["Id", "Name", "Category", "Description", "Act", "Rules", "Form",
           "Section Description", "Periodicity Due Date",
           "Consequence", "City: Exception", "Sector: Exception",
           "Business Activity Exception", "Sector Activity Exception",
           "Employee Factor: Exception", "Help Text", "Prepared by", "Comment",
           "Conditions (indicates logic e.g. AND, OR)"]

# Query Template
BQUERY = {"bool": {}}
QUERY = {"query": {"constant_score": {"filter": BQUERY}},
         "sort": [{"metadata.Id": {"order": "asc", "unmapped_type": "long"}}]
        }


def create_index(es, index):
    """
    Create Index with appropriate settings.
    """
    settings = {"analysis": {"analyzer": {
                                    "my_analyzer": {
                                        "type": "custom",
                                        "tokenizer": "my_tokenizer",
                                        "filter": ["trim"]
                                        }
                                    },
                                   "tokenizer": {
                                        "my_tokenizer": {
                                            "type": "keyword"
                                        }
                                   }
                                }
                 }
    mappings = {"activity": {"dynamic_templates": [
                                {"attributes": {
                                    "match_mapping_type": "string",
                                    "path_match": "attributes.*",
                                    "mapping": {
                                        "type": "keyword",
                                        "analyzer": "my_analyzer",
                                        "search_analyzer": "my_analyzer"
                                    }
                                }},
                                {"metadata": {
                                    "match_mapping_type": "text",
                                    "path_match": "metadata.*",
                                    "mapping": {
                                        "type": "text",
                                        "index": False,
                                        "enabled": False,
                                    }
                                }}
                            ]}
                }
    body = {
              "settings": settings,
              "mappings": mappings
            }
    es.indices.create(index=index, ignore=400, body=body)
    print "Created Index %s!" % index
    return


def delete_index(es, index):
    """
    Delete index.
    """
    es.indices.flush(index=index)
    es.indices.delete(index=index, ignore=[400, 404])
    es.indices.refresh(index=index)
    print "Deleted Index %s!" % index
    return


def extract_headers(row, search=False):
    """
    Extract Column headers from master data csv.
    """
    global HEADERS, ATTRIBUTES
    while row and row[-1] is "":
        row.pop()
    if not search:
        for head in row:
            if head.value:
                value = head.value.strip()
                HEADERS.append(value)
                if value not in AFIELDS:
                    ATTRIBUTES.append(head.value)
        HEADERS.insert(0, "Id")
        return
    else:
        return row


def process_master_data_record(row):
    """
    Process Master Data Record.
    """
    global HEADERS
    global DATA
    row_dict = OrderedDict(zip(HEADERS, row))
    attributes = OrderedDict()
    metadata = OrderedDict()
    for key, value in row_dict.items():
        if value is None:
            value = ""
        valid_key = key.strip()
        if isinstance(value, basestring):
            value = value.strip()
            value = value.replace('\n',' ').replace('\r',' ')
            value = value.replace('  ',' ')
        if valid_key not in AFIELDS:
            attributes[valid_key] = value.replace("; ", ";").split(";")
        else:
            metadata[valid_key] = value
    data_dict = {'attributes': attributes, 'metadata': metadata}
    DATA.append(data_dict)
    return


def extract_master_data(master_data):
    """
    Read CSV and index activities.
    """
    st = time.time()
    global DATA
    counter = 0
    wb = load_workbook(filename = master_data)
    # wb = load_workbook(filename=master_data, read_only=True, data_only=True)
    ws = wb.get_active_sheet()
    for row in ws.rows:
        if counter == 0:
            extract_headers(row)
        else:
            if not any(cell.value for cell in row):
                pass
            else:
                values = [cell.value for cell in row]
                values.insert(0, counter)
                process_master_data_record(values)
        counter += 1
    print "Total Records %d in %f" % (len(DATA), time.time()-st)
    return


def index_master_data(es, index, doctype):
    """
    Index master data.
    """
    st = time.time()
    print "Indexing Data..."
    global DATA
    counter = 1
    for record in DATA:
        try:
            result = es.index(index=index, doc_type=doctype, id=counter,
                          body=record, refresh=True)
        except Exception,e:
            print str(e)
            print record
            sys.exit(-1)
        counter+=1
    es.indices.refresh(index=index)
    print "Indexed master data in %f" % (time.time() - st)
    return


def get_nested_query(skey, svalue):
    """
    Get Nested query.
    """
    global BQUERY
    local_query = deepcopy(BQUERY)
    val_list = svalue.split(";")
    qlist = []
    for val in val_list:
        temp = {"term": {"attributes." + skey: val.strip()}}
        qlist.append(temp)
    local_query["bool"]["should"] = qlist
    return local_query


def process_query(qdict):
    """
    Prepare Query
    """
    global QUERY
    must_list = []
    must_not_list = []

    for key, value in qdict["attributes"].items():
        exception_exists = False
        if key.find("Exception") >= 0:
            exception_exists = True
            key = key.replace("Exception", "").strip()
        if value not in ["", None]:
            if value.find(";") >= 0:
                temp = get_nested_query(key, value)
            else:
                temp = {"term": {"attributes." + key: value.strip()}}
        if exception_exists:
            must_not_list.append(temp)
        else:
            must_list.append(temp)
    QUERY.get("query", {}).get("constant_score", {}).get("filter", {}).get("bool", {})["must_not"] = must_not_list
    QUERY.get("query", {}).get("constant_score", {}).get("filter", {}).get("bool", {})["must"] = must_list
    return QUERY


def process_search_data_record(row, headers):
    """
    Process Search data record.
    """
    row_dict = OrderedDict(zip(headers, row))
    attributes = OrderedDict()
    metadata = OrderedDict()
    for key, value in row_dict.items():
        if value not in ["", None]:
            valid_key = key.strip()
            if isinstance(value, basestring):
                value = value.strip()
            if valid_key not in AFIELDS:
                attributes[valid_key] = value
            else:
                metadata[valid_key] = value
    data_dict = {'attributes': attributes, 'metadata': metadata}
    return data_dict


def get_search_query(search_input):
    """
    Get Search Query.
    """
    # Set query dict parameters
    counter = 0
    with open(search_input, "rb") as csvfile:
        reader = csv.reader(csvfile, delimiter=',')
        for row in reader:
            if counter == 0:
                headers = extract_headers(row, search=True)
            else:
                query_dict = process_search_data_record(row, headers)
            counter += 1
        try:
            next(csvfile)
        except Exception, e:
            print "End of file!"
    query = process_query(query_dict)
    print"Query = = %s"%(query)
    return query


def search_index(es, index, doctype, search_input):
    """
    Search Index.
    """
    st = time.time()
    global HEADERS

    # Get Search Query
    query_dict = get_search_query(search_input)
    # query_dict = {"sort": [{"metadata.Id": {"unmapped_type": "long", "order": "asc"}}], "query": {"constant_score": {"filter": {"bool": {"must_not": [], "must": [{"term": {"attributes.Sector": "ALL"}}, {"term": {"attributes.City": "ALL"}}, {"term": {"attributes.Product": "ALL"}}, {"term": {"attributes.Financial Factor": "ALL"}}, {"term": {"attributes.Assets": "ALL"}}, {"term": {"attributes.Zone": "ALL"}}, {"term": {"attributes.Country": "India"}}, {"term": {"attributes.Employee Strength": "ALL"}}, {"term": {"attributes.Tax Factor": "ALL"}}, {"term": {"attributes.Employee Factor": "ALL"}}, {"term": {"attributes.Business Activity": "ALL"}}, {"term": {"attributes.State": "ALL"}}, {"term": {"attributes.Environment Factor": "ALL"}}, {"term": {"attributes.Sector Activity": "ALL"}}, {"term": {"attributes.Establishment Type": "ALL"}}, {"term": {"attributes.Entity Type": "ALL"}}, {"term": {"attributes.Ownership": "ALL"}}, {"term": {"attributes.Consumable": "ALL"}}, {"term": {"attributes.Wage Factor": "ALL"}}, {"term": {"attributes.Event Factor": "ALL"}}]}}}}}
    query = json.dumps(query_dict)
    print query
    # Search using query
    result = es.search(index=index, doc_type=doctype,
                       body=query, size=4000)

    print "Total Hits %d in %f" % (len(result["hits"]['hits']), time.time() - st)

    # Write the result to a xlsx
    wb = Workbook()
    ws = wb.active
    counter = 1
    for head in HEADERS:
        _=ws.cell(row=1, column=counter, value=head)
        counter += 1
    row = 2
    for hit in result['hits']['hits']:
        cell = 1
        for header in HEADERS:
            value = hit["_source"]['attributes'].get(header, hit["_source"]['metadata'].get(header, ""))
            if isinstance(value, basestring):
                value = value.encode("utf-8")
            value =str(value)
            ws.cell(column=cell, row=row, value=value)
            cell += 1
        row += 1
    filename = search_input[:-4] + ".xlsx"
    print "\n\n*****************"
    print "INFO: Applicable activities written to [%s]\n\n*****************" % filename
    wb.save(filename=filename)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('-i', '--input', help='Master Data CSV.',
                        default="master_data.xlsx")
    parser.add_argument('-s', '--search', help='Search Data CSV.',
                        default="search_query.csv")
    parser.add_argument('-o', '--operation', help='operation to be performed create/search',
                        default="")

    args = parser.parse_args()

    operation = args.operation.lower()
    if operation not in ['create', 'search']:
        print "ERROR: Invalid argument value ['%s'] for operation, enter 'create' or 'search'" % operation

    # Extract and process master data
    extract_master_data(args.input)

    # Index Parameters
    index = "master_data"
    doctype = "activity"

    # Elastic Search Instance
    es = Elasticsearch(settings.elastic_url)

    if operation == 'create':
        # Delete all indices
        print "INFO: Deleting previous indexes"
        es.indices.delete(index="_all")
        es.indices.refresh()
        print "INFO: Deleted previous indexes"
        # Index Management
        print "INFO: creating new indexes"
        create_index(es, index)
        index_master_data(es, index, doctype)
        print "INFO: Indexing completed"

    elif operation == 'search':
        # Search terms
        print "INFO: Generating applicable activities"
        search_index(es, index, doctype, args.search)
        print "INFO: Generation completed"
    print "Done!"

if __name__ == '__main__':
    main()
