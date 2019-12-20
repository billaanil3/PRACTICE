import xlrd
from openpyxl import Workbook


def _write_excel_data(wb, headers, data):
    """
    writing categories,locations,parties,users,privilegs,locaton_category_users data
    multiple=True ,if multiple cells of data
    """
    ws = wb.create_sheet("Test Data", 0)
    _write_headers(ws, headers)
    _write_to_multiple_cell(ws, data)


def _write_to_multiple_cell(ws, sheet_data):
    """
    writing data to
    rows = 1,2,3...........
    cell = 1,2,3...........
    """
    print "***************", sheet_data
    row = 2
    for value in sheet_data:
        cell = 1
        for val in value:
            ws.cell(column=cell, row=row, value=val)
            cell += 1
        row += 1


def _write_headers(ws, headers):
    """
    writing headers
    """
    counter = 1
    for head in headers:
        _ = ws.cell(row=1, column=counter, value=head)
        counter += 1


file_path = '/home/anil/Desktop/checklist_sample.xlsx'
new_file_path = '/home/anil/Desktop/reddy_sample.xlsx'
book = xlrd.open_workbook(file_path)
headers = ["Transaction_Name", "Abbreviation", "Document_Name"]
main_data = []
for sheet in book.sheets():
    with open(new_file_path, "w") as file:
        for rownum in range(1, sheet.nrows):
            val1 = sheet.row(rownum)[0].value.strip()
            val2 = sheet.row(rownum)[1].value.strip()
            val3 = sheet.row(rownum)[2].value.strip()
            data = [val1, val2, val3]
            main_data.append(data)
        wb = Workbook()
        # print "---------------", data
        _write_excel_data(wb, headers, main_data)
        wb.save(new_file_path)
