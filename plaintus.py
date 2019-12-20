from __future__ import print_function
import openpyxl 

# wb=openpyxl.load_workbook(path)
workbook = openpyxl.load_workbook("plaintus.xlsx",read_only=True)
# print (workbook)
sheet = workbook.active
# sheet = workbook.get_sheet_by_name('Onboard')
# print sheet
print ("--------Title-------->",sheet.title)
print ("--------Rows--------->",sheet.max_row)
print ("--------Columns------>",sheet.max_column)
# a = sheet['A2']
# print a.value
# cells = sheet['A2':'M2']
# for cell_value in cells:
#     print cell_value.value
# workbook.save()
#****************************************************************
# for row in sheet.iter_rows(min_row=2, min_col=1, max_row=10, max_col=13):
#     for cell in row:
#         print(cell.value)
#     print"*****************************"
    
#*****************************************************************
# rows = 10
# columns = 13
# values = []
# for row in range(1,rows+1):
#     for col in range(1,columns+1):
#         d = sheet.cell(row=row,column=col)
#         # print (values.append(str(d.value)))
#         # print (type(str(d.value)))
#         print ((d.value))
#         # print ('%-8s'%d.value,end=" ")
#     print ("*********************")

#**********************************************************************
headers = ['Location', 'Incident_type', 'Agency', 'Status', 'Risk', 'Receipt_date', 'Recipient_type', 'Recipient', 'Receipt_mode', 'Description Consequence','Additional_fields', 'Incident_file']
# l1 = ''
# l2 = ''
for row in sheet.iter_rows():
    values = []
    for cell in row:
        values.append(cell.value)
    print(zip(headers, values))
    # db.dispute.insert(location=values)
    # print ("$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$")
