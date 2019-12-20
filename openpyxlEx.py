# Create a new file
'''
from openpyxl import Workbook
import time

book = Workbook()
sheet = book.active

sheet['A1'] = 'Anil'
sheet['A2'] = 'Billa'
sheet['A3'] = 25
now = time.strftime("%x")
sheet['A4'] = now
book.save("openpyxl.xlsx")
'''

# Writing to a cell
'''
from openpyxl import Workbook
book = Workbook()
sheet = book.active
sheet['A1'] = 143
sheet.cell(row=2,column=2).value = 'Anil'
book.save('openpyxl2.xlsx')
'''
# Appending Values
from openpyxl import Workbook
book = Workbook()
sheet = book.active
rows = (
    (88,46,57),
    (89,38,12),
    (23, 59, 78),
    (56, 21, 98),
    (24, 18, 43),
    (34, 15, 67)
)
for row in rows:
    sheet.append(row)
#Reading Multiple Values
print sheet['C5'].value   # 1st way to get cell value
print sheet.cell(2,3).value # 2nd way to get cell value
a1 = sheet['A1']
a2 = sheet['A2']
a3 = sheet.cell(row=2,column=2)
print a1.value
print a2.value
print a3.value
#Reading Multiple Values
cells = sheet['A1':'B6']
# print type(cells)
for c1,c2 in cells:
    # print c1
    # print c2
    print ("{0:4} {1:4}".format(c1.value,c2.value))
book.save('openpyxl3.xlsx')